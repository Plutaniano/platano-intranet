from models import db, Investimentos, Previdencia, BancoXP, CoCorretagem, IncentivoPrevidencia
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import datetime
import re
from parse import *

def parse_excel(filename: str) -> None:
    wb = load_workbook(filename)
    date = get_date(wb['Investimentos'].cell(1, 1).value)
    gerado = get_datetime(wb['Investimentos'].cell(1, 4).value)

    parse_investimentos(wb['Investimentos'], date)
    parse_previdencia(wb['Previdência'], date)
    parse_co_corretagem(wb['Co-Corretagem'], date)
    parse_incentivo_previdencia(wb['Incentivo Previdência'], date)
    parse_banco_xp(wb['Banco XP'], date)



def parse_investimentos(ws: Worksheet, ano_mes: datetime.date) -> None:
    for row in ws.iter_rows():
        if row[0].value not in ['AJUSTES', 'RECEITAS']:
            continue
        
        else:
            entry = Investimentos(ano_mes=ano_mes,
                                  classificacao=row[0].value,
                                  produto=row[1].value,
                                  nivel1=row[2].value,
                                  nivel2=row[3].value,
                                  nivel3=row[4].value,
                                  cliente=int(row[5].value),
                                  master=int(row[6].value[1:]),
                                  data=row[7].value.date(),
                                  receita_bruta=int(100*row[8].value),
                                  receita_liquida=int(100*row[9].value),
                                  comissao_escritorio_porcento=float(0.01*row[10].value),
                                  comissao_escritorio=int(100*row[11].value),
                                  
                                  codigo_assessor=cell_to_assessor_int(row[13]),
                                  assessor_direto_comissao_porcento=float(0.01*row[14].value),
                                  assessor_direto_comissao=int(100*row[15].value),
                                  
                                  assessor_indireto1_codigo=cell_to_assessor_int(row[17]),
                                  assessor_indireto1_comissao_porcento=float(0.01*row[18].value),
                                  assessor_indireto1_comissao=int(100*row[19].value),
                                  
                                  assessor_indireto2_codigo=cell_to_assessor_int(row[21]),
                                  assessor_indireto2_comissao_porcento=float(0.01*row[22].value),
                                  assessor_indireto2_comissao=int(100*row[23].value),
                                  
                                  assessor_indireto3_codigo=cell_to_assessor_int(row[25]),
                                  assessor_indireto3_comissao_porcento=float(0.01*row[26].value),
                                  assessor_indireto3_comissao=int(100*row[27].value)
                                  )
            print(entry)
            db.session.add(entry)
            db.session.commit()



def parse_previdencia(ws: Worksheet, ano_mes: datetime.date) -> None:
    for row in ws.iter_rows():
        if row[2].value not in ['PLATANO AGENTE AUTONOMO DE INVESTIMENTOS LTDA', 'Platano Investimentos']:
            continue
        
        else:
            entry = Previdencia(ano_mes=ano_mes,
                                tipo=row[0].value,
                                competencia=row[1].value.date(),
                                parceiro=row[2].value,
                                codigo_assessor=row[3].value,
                                certificado=row[4].value if isinstance(row[4].value, int) else 0,
                                cpf=str(row[5].value),
                                codigo_cliente=row[6].value if isinstance(row[6].value, int) else 0,
                                up=str(row[7].value),
                                
                                # Dados do produto
                                seguradora=str(row[8].value),
                                produto=str(row[9].value),
                                data_emissao=row[10].value.date() if isinstance(row[10].value, datetime.date) else None,
                                reserva=int(100 * row[11].value if isinstance(row[11].value, float) else 0),
                                tx_adm=float(row[12].value if isinstance(row[12].value, float) else 0),
                                
                                # TAF
                                taf_base=int(100 * row[13].value if isinstance(row[13].value, float) else 0),
                                taf_repasse_porcento=float(row[14].value if isinstance(row[14].value, float) else 0),
                                taf_receita=int(row[15].value if isinstance(row[15].value, int) else 0),
                                
                                # 1a Aplicação Mensal
                                primeira_aplicacao_mensal_base=int(100 * row[16].value if isinstance(row[16].value, float) else 0),
                                primeira_aplicacao_mensal_repasse=float(row[17].value if isinstance(row[17].value, float) else 0),
                                primeira_aplicacao_mensal_receita=int(row[18].value if isinstance(row[18].value, int) else 0),

                                # Aportes/Prêmio
                                aportes_base=int(row[19].value * 100 if isinstance(row[19].value, float) else 0),
                                aportes_repasse_porcento=float(row[20].value if isinstance(row[20].value, float) else 0),
                                aportes_receita=int(100 * row[21].value if isinstance(row[21].value, int) else 0),

                                # Portabilidade
                                portabilidade_base=int(100 * row[22].value if isinstance(row[22].value, float) else 0),
                                portabilidade_repasse_porcento=float(row[23].value if isinstance(row[23].value, float) else 0),
                                portabilidade_receita=int(100 * row[24].value if isinstance(row[24].value, int) else 0),

                                # Receita
                                receita_bruta_total=int(100 * row[25].value),
                                ir_sobre_receita_bruta=int(100 * row[26].value),
                                receita_liquida_total=int(100 * row[27].value),
                                obs=str(row[28].value)
                                )
            print(entry)
            db.session.add(entry)
            db.session.commit()


def parse_co_corretagem(ws: Worksheet, ano_mes: datetime.date) -> None:
    for row in ws.iter_rows():
        if not isinstance(row[1].value, datetime.date):
            continue
        
        else:
            entry = CoCorretagem(ano_mes=ano_mes,
                                tipo=row[0].value,
                                competencia=row[1].value.date(),
                                parceiro=row[2].value,
                                codigo_assessor=row[3].value,
                                certificado=row[4].value if isinstance(row[4].value, int) else -0,
                                cpf=str(row[5].value),
                                codigo_cliente=row[6].value if isinstance(row[6].value, int) else 0,
                                nome_cliente=str(row[7].value),
                                
                                # Dados do produto
                                seguradora=str(row[8].value),
                                produto=str(row[9].value),
                                data_emissao=row[10].value.date() if isinstance(row[10].value, datetime.date) else datetime.date(1970, 1, 1),
                                reserva=int(100 * row[11].value if isinstance(row[11].value, float) else 0),
                                tx_adm=float(row[12].value if isinstance(row[12].value, float) else 0),
                                
                                # TAF
                                taf_base=int(100 * row[13].value if isinstance(row[13].value, float) else 0),
                                taf_repasse_porcento=float(row[14].value if isinstance(row[14].value, float) else 0),
                                taf_receita=int(row[15].value if isinstance(row[15].value, int) else 0),
                                
                                # 1a Aplicação Mensal
                                primeira_aplicacao_mensal_base=int(100 * row[16].value if isinstance(row[16].value, float) else 0),
                                primeira_aplicacao_mensal_repasse=float(row[17].value if isinstance(row[17].value, float) else 0),
                                primeira_aplicacao_mensal_receita=int(row[18].value if isinstance(row[18].value, int) else 0),

                                # Aportes/Prêmio
                                aportes_base=int(row[19].value * 100 if isinstance(row[19].value, float) else 0),
                                aportes_repasse_porcento=float(row[20].value if isinstance(row[20].value, float) else 0),
                                aportes_receita=int(100 * row[21].value if isinstance(row[21].value, int) else 0),

                                # Portabilidade
                                portabilidade_base=int(100 * row[22].value if isinstance(row[22].value, float) else 0),
                                portabilidade_repasse_porcento=float(row[23].value if isinstance(row[23].value, float) else 0),
                                portabilidade_receita=int(100 * row[24].value if isinstance(row[24].value, int) else 0),

                                # Receita
                                receita_total=int(100 * row[25].value),
                                obs=str(row[26].value)
                                )
            print(entry)
            db.session.add(entry)
            db.session.commit()



def parse_incentivo_previdencia(ws: Worksheet, ano_mes: datetime.date) -> None:
    for row in ws.iter_rows():
        if not isinstance(row[0].value, datetime.date):
            continue
        
        else:
            entry = IncentivoPrevidencia(
                                ano_mes=ano_mes,
                                mes_referencia=row[0].value.date(),
                                status_docusign=row[1] == 'Assinado',
                                codigo_escritorio=row[2].value,
                                escritorio=row[3].value,
                                codigo_assessor=row[4].value,
                                codigo_cliente=row[5].value,
                                certificado=row[6].value,
                                movimentacao_cliente=int(100 * row[7].value),
                                adiantamento_previdencia=int(100 * row[8].value)
                                )
            print(entry)
            db.session.add(entry)
            db.session.commit()

def parse_banco_xp(ws: Worksheet, ano_mes: datetime.date) -> None:
    for row in ws.iter_rows():
        if not isinstance(row[0].value, datetime.date):
            continue
        
        else:
            entry = BancoXP(
                                ano_mes=ano_mes,
                                competencia=row[0].value.date(),
                                codigo_escritorio=row[1].value,
                                parceiro=row[2].value,
                                codigo_assessor=row[3].value if isinstance(row[3].value, int) else str(row[3].value)[1:],
                                operacao=row[4].value,
                                codigo_cliente=row[5].value,
                                produto=row[6].value,
                                data_contratacao=row[7].value,
                                data_vencimento=row[8].value,
                                valor_contratado=int(100 * row[9].value),
                                juros_aa=float(row[10].value),
                                comissao_escritorio_porcento_aa=float(row[11].value),
                                comissao_atualizada_acumulada=int(100 * row[12].value),
                                deducoes=float(row[13].value),
                                total_receita=int(100 * row[14].value)
                                )
            print(entry)
            db.session.add(entry)
            db.session.commit()

def get_date(s: str) -> datetime.date:
    m = re.findall('(\d{2})\-(\d{4})', s)[0]
    return datetime.date(int(m[1]), int(m[0]), 1)

def get_datetime(s: str) -> datetime.datetime:
    s = s[8:]
    return datetime.datetime.strptime(s, '%d/%m/%Y - %H:%M')

def cell_to_assessor_int(cell) -> int:
    if cell.value == '':
        return 0
    
    if cell.value == 'PAN':
        return -1

    else:
        return int(cell.value.replace('A', ''))