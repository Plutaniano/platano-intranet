from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .investimentos import parse_investimentos
from .previdencia import parse_previdencia
from .cocorretagem import parse_co_corretagem
from .cambio import parse_cambio
from .bancoxp import parse_banco_xp

from ..models.investimentos import Investimentos
from ..models.previdencia import Previdencia
from ..models.cocorretagem import CoCorretagem
from ..models.cambio import Cambio
from ..models.bancoxp import BancoXP

from typing import List, Tuple
import datetime
import re

def parse_excel(filename: str, mes_de_entrada: datetime.date) -> List[Tuple[object, bool]]:
    wb = load_workbook(filename)

    tabelas = [
        ['Investimentos', parse_investimentos],
        ['Previdência', parse_previdencia],
        ['Co-Corretagem', parse_co_corretagem],
        ['Banco XP', parse_banco_xp],
        ['Câmbio', parse_cambio]
    ]
    
    resultados = []

    for i, (nome, parser) in enumerate(tabelas):
        if nome in wb.sheetnames:
            ws = wb[nome]
            parser(ws, mes_de_entrada)
            resultados.append((nome, True))
        else:
            print(f'Tabela "{nome}" não encontrada.')
            resultados.append((nome, False))
    
    return resultados
        

def get_date(s: str) -> datetime.date:
    m = re.findall('(\d{2})\-(\d{4})', s)[0]
    return datetime.date(int(m[1]), int(m[0]), 1)

def get_datetime(s: str) -> datetime.datetime:
    s = s[8:]
    return datetime.datetime.strptime(s, '%d/%m/%Y - %H:%M')

